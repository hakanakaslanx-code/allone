"""Automated mapper for Wayfair's "Product Addition Template" workbooks.

This module reads a local master data file together with the official Wayfair
template, discovers the template columns by prefix, applies fuzzy matching to
map fields, normalizes values against the "Valid Values" sheet, and writes the
result into a copy of the template while keeping the original formatting.
"""

from __future__ import annotations

import argparse
import difflib
import logging
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Logging helpers
# ---------------------------------------------------------------------------


def setup_logger(log_path: Path) -> logging.Logger:
    """Create a logger that writes verbose information to ``log_path``."""

    logger = logging.getLogger("wayfair_mapper")
    logger.setLevel(logging.INFO)
    logger.propagate = False

    if not logger.handlers:
        handler = logging.FileHandler(log_path, encoding="utf-8")
        formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
        handler.setFormatter(formatter)
        logger.addHandler(handler)

    return logger


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------


def _clean_value(value: object) -> str:
    """Return a trimmed string representation of ``value`` suitable for Excel."""

    if value is None:
        return ""

    if isinstance(value, float) and pd.isna(value):
        return ""

    if isinstance(value, str):
        return value.strip()

    return str(value).strip()


def _split_semicolon(value: str, limit: Optional[int] = None) -> List[str]:
    """Split ``value`` by semicolon into trimmed items respecting ``limit``."""

    if not value:
        return []

    parts = [segment.strip() for segment in value.split(";")]
    cleaned = [segment for segment in parts if segment]
    if limit is None:
        return cleaned
    return cleaned[:limit]


def _normalize_header_name(name: str) -> str:
    """Return a simplified representation of ``name`` for comparisons."""

    return re.sub(r"[^a-z0-9]", "", name.lower())


def _compute_match_score(column_name: str, synonyms: Sequence[str]) -> float:
    """Return the best similarity score between ``column_name`` and ``synonyms``."""

    normalized_column = _normalize_header_name(column_name)
    best_score = 0.0

    for synonym in synonyms:
        normalized_synonym = _normalize_header_name(synonym)
        if not normalized_synonym:
            continue
        if normalized_column == normalized_synonym:
            return 1.0
        score = difflib.SequenceMatcher(None, normalized_column, normalized_synonym).ratio()
        best_score = max(best_score, score)

    return best_score


@dataclass
class FieldDefinition:
    """Describe a logical mapping between master data and template fields."""

    canonical: str
    template_targets: List[str]
    synonyms: Sequence[str]
    description: str
    valid_key: Optional[str] = None
    split_semicolon: bool = False
    max_items: Optional[int] = None
    multi_source: bool = False

    def __post_init__(self) -> None:
        if self.max_items is None:
            self.max_items = len(self.template_targets)


FIELD_DEFINITIONS: List[FieldDefinition] = [
    FieldDefinition(
        canonical="core::supplierPartNumber",
        template_targets=["core::supplierPartNumber"],
        synonyms=["sku", "rugno", "rug no", "supplier part number", "item number"],
        description="SKU / RugNo",
    ),
    FieldDefinition(
        canonical="core::universalProductCode",
        template_targets=["core::universalProductCode"],
        synonyms=["upc", "barcode", "bar code", "ean"],
        description="UPC / Barcode",
    ),
    FieldDefinition(
        canonical="core::productName",
        template_targets=["core::productName"],
        synonyms=["product name", "name", "title"],
        description="Product Name / Title",
    ),
    FieldDefinition(
        canonical="core::collectionName",
        template_targets=["core::collectionName"],
        synonyms=["collection", "collection name"],
        description="Collection",
    ),
    FieldDefinition(
        canonical="price::wholesalePrice",
        template_targets=["price::wholesalePrice"],
        synonyms=["wholesale", "wholesale price", "base", "base price"],
        description="Wholesale / Base",
    ),
    FieldDefinition(
        canonical="price::minimumAdvertizedPrice",
        template_targets=["price::minimumAdvertizedPrice"],
        synonyms=["map", "minimum advertised price"],
        description="MAP",
    ),
    FieldDefinition(
        canonical="price::manufacturerSuggestedRetailPrice",
        template_targets=["price::manufacturerSuggestedRetailPrice"],
        synonyms=["msrp", "retail price", "suggested retail"],
        description="MSRP",
    ),
    FieldDefinition(
        canonical="featureDescription::romanceCopy",
        template_targets=["featureDescription::romanceCopy"],
        synonyms=["description", "longdesc", "long description", "product description"],
        description="Description / LongDesc",
    ),
    FieldDefinition(
        canonical="featureDescription::genericFeatures",
        template_targets=[
            "featureDescription::genericFeatures",
            "featureDescription::genericFeatures.1",
            "featureDescription::genericFeatures.2",
            "featureDescription::genericFeatures.3",
            "featureDescription::genericFeatures.4",
        ],
        synonyms=["feature bullet", "features", "bullet"],
        description="Feature Bullet 1-5",
        split_semicolon=True,
        multi_source=True,
        max_items=5,
    ),
    FieldDefinition(
        canonical="shippingAndFulfillment::weight",
        template_targets=["shippingAndFulfillment::weight"],
        synonyms=["weight", "package weight"],
        description="Package Weight",
    ),
    FieldDefinition(
        canonical="shippingAndFulfillment::height",
        template_targets=["shippingAndFulfillment::height"],
        synonyms=["height", "package height"],
        description="Package Height",
    ),
    FieldDefinition(
        canonical="shippingAndFulfillment::width",
        template_targets=["shippingAndFulfillment::width"],
        synonyms=["width", "package width"],
        description="Package Width",
    ),
    FieldDefinition(
        canonical="shippingAndFulfillment::depth",
        template_targets=["shippingAndFulfillment::depth"],
        synonyms=["depth", "package depth", "length"],
        description="Package Depth",
    ),
    FieldDefinition(
        canonical="media::imageUrls",
        template_targets=[
            "media::image1",
            "media::image2",
            "media::image3",
            "media::image4",
            "media::image5",
        ],
        synonyms=["imageurls", "image url", "image urls", "images", "image links"],
        description="Image URLs",
        split_semicolon=True,
        max_items=5,
    ),
    FieldDefinition(
        canonical="core::countryOfManufacturer",
        template_targets=["core::countryOfManufacturer"],
        synonyms=["country", "country of origin", "country of manufacturer", "made in"],
        description="Country",
    ),
    FieldDefinition(
        canonical="attr::shape",
        template_targets=["attr::shape"],
        synonyms=["shape", "rug shape"],
        description="Rug Shape",
        valid_key="shape",
    ),
    FieldDefinition(
        canonical="attr::material",
        template_targets=["attr::material"],
        synonyms=["material", "rug material", "fabric", "construction"],
        description="Rug Material",
        valid_key="material",
    ),
    FieldDefinition(
        canonical="attr::technique",
        template_targets=["attr::technique"],
        synonyms=["technique", "rug technique", "weave"],
        description="Rug Technique",
        valid_key="technique",
    ),
    FieldDefinition(
        canonical="attr::color",
        template_targets=["attr::color"],
        synonyms=["color", "colour", "primary color"],
        description="Color",
        valid_key="color",
    ),
    FieldDefinition(
        canonical="attr::careInstructions",
        template_targets=["attr::careInstructions"],
        synonyms=["care", "care instructions", "cleaning"],
        description="Care",
        valid_key="care",
    ),
]

PROP65_TARGET = "propSixtyFive::warningRequired"
AUTO_ACCEPT_THRESHOLD = 0.85
PROMPT_THRESHOLD = 0.6
MULTI_ACCEPT_THRESHOLD = 0.7


# ---------------------------------------------------------------------------
# Template helpers
# ---------------------------------------------------------------------------


def _find_sheet(workbook, target_name: str) -> Optional[Worksheet]:
    """Return the worksheet that best matches ``target_name`` (case insensitive)."""

    lower_target = target_name.lower()
    for sheet_name in workbook.sheetnames:
        if sheet_name.lower() == lower_target:
            return workbook[sheet_name]

    closest = difflib.get_close_matches(lower_target, [s.lower() for s in workbook.sheetnames], n=1, cutoff=0.6)
    if closest:
        index = [s.lower() for s in workbook.sheetnames].index(closest[0])
        return workbook[workbook.sheetnames[index]]

    return None


def detect_template_columns(sheet: Worksheet) -> Dict[str, int]:
    """Discover template columns by reading the header row of ``sheet``."""

    headers: Dict[str, int] = {}

    try:
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return headers

    for idx, raw_header in enumerate(first_row, start=1):
        if raw_header is None:
            continue
        header = str(raw_header).strip()
        if not header:
            continue
        headers[header.lower()] = idx

    return headers


def _match_column(headers: Mapping[str, int], target: str) -> Optional[int]:
    """Return the column index for ``target`` using case-insensitive fuzzy matching."""

    normalized_target = target.lower()

    if normalized_target in headers:
        return headers[normalized_target]

    if "::" in normalized_target:
        prefix, suffix = normalized_target.split("::", 1)
        suffix_matches = [name for name in headers if name.endswith(suffix)]
        if suffix_matches:
            best_suffix = difflib.get_close_matches(normalized_target, suffix_matches, n=1, cutoff=0.6)
            if best_suffix:
                return headers[best_suffix[0]]

        suffix_matches = [name for name in headers if name.endswith(suffix.replace(".", ""))]
        if suffix_matches:
            best_suffix = difflib.get_close_matches(normalized_target, suffix_matches, n=1, cutoff=0.6)
            if best_suffix:
                return headers[best_suffix[0]]

    best = difflib.get_close_matches(normalized_target, list(headers.keys()), n=1, cutoff=0.6)
    if best:
        return headers[best[0]]

    return None


# ---------------------------------------------------------------------------
# Valid value normalization
# ---------------------------------------------------------------------------


def normalize_with_valid_values(
    field_name: str,
    value: str,
    valid_map: Mapping[str, Iterable[str]],
    logger: logging.Logger,
    alias: Optional[str] = None,
) -> Tuple[str, bool, bool, bool]:
    """Normalize ``value`` according to the ``Valid Values`` sheet."""

    cleaned_value = _clean_value(value)
    had_input = bool(cleaned_value)
    if not cleaned_value:
        return "", False, True, False

    candidates: List[str] = []
    if alias:
        candidates.append(alias.lower())
    if "::" in field_name:
        candidates.append(field_name.split("::", 1)[-1].lower())
    candidates.append(field_name.lower())

    valid_values: Optional[Iterable[str]] = None
    for key in candidates:
        if key in valid_map:
            valid_values = valid_map[key]
            break

    if not valid_values:
        return cleaned_value, False, True, True

    normalized_map = {str(item).strip().lower(): str(item).strip() for item in valid_values if str(item).strip()}

    if cleaned_value.lower() in normalized_map:
        normalized = normalized_map[cleaned_value.lower()]
        return normalized, normalized != cleaned_value, True, True

    candidates = difflib.get_close_matches(cleaned_value.lower(), list(normalized_map.keys()), n=1, cutoff=0.75)
    if candidates:
        return normalized_map[candidates[0]], True, True, True

    logger.info("Value '%s' for field '%s' does not match valid options", cleaned_value, field_name)
    return "", False, False, True


def _load_valid_values(workbook) -> Dict[str, List[str]]:
    """Read the ``Valid Values`` sheet and build a dictionary for normalization."""

    sheet = _find_sheet(workbook, "Valid Values")
    if sheet is None:
        return {}

    data = list(sheet.iter_rows(values_only=True))
    if not data:
        return {}

    valid_map: Dict[str, List[str]] = {}

    for row in data:
        if not row:
            continue
        field_name = _clean_value(row[0])
        if not field_name:
            continue
        values = [_clean_value(item) for item in row[1:] if _clean_value(item)]
        if values:
            valid_map[field_name.lower()] = values

    return valid_map


# ---------------------------------------------------------------------------
# Mapping persistence helpers
# ---------------------------------------------------------------------------


def _load_saved_mappings(path: Optional[Path]) -> Dict[str, List[str]]:
    if path is None or not path.exists():
        return {}

    try:
        with path.open("r", encoding="utf-8") as handle:
            raw = yaml.safe_load(handle) or {}
    except Exception:
        return {}

    mappings: Dict[str, List[str]] = {}
    for key, value in raw.items():
        if isinstance(value, str):
            mappings[key] = [value]
        elif isinstance(value, Sequence):
            mappings[key] = [str(item) for item in value]
    return mappings


def _save_mappings(path: Optional[Path], mappings: Mapping[str, List[str]]) -> None:
    if path is None:
        return

    serializable = {key: value if len(value) != 1 else value[0] for key, value in mappings.items() if value}

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        yaml.safe_dump(serializable, handle, allow_unicode=True, sort_keys=True)


def _prompt_for_selection(
    definition: FieldDefinition,
    candidates: Sequence[Tuple[float, str]],
    allow_multiple: bool,
) -> List[str]:
    """Prompt the user to select columns for ``definition``."""

    print()
    print(f"{definition.description} için kolon seçin (0 = atla):")
    for idx, (score, name) in enumerate(candidates, start=1):
        print(f"  {idx}) {name} [skor: {score:.2f}]")

    selection: List[str] = []

    while True:
        raw = input("Seçiminiz: ").strip()
        if not raw:
            print("Lütfen bir değer girin veya 0 yazarak atlayın.")
            continue
        if raw == "0":
            break
        if allow_multiple:
            parts = [part.strip() for part in raw.split(",") if part.strip()]
            chosen: List[str] = []
            valid = True
            for part in parts:
                if not part.isdigit():
                    valid = False
                    break
                index = int(part)
                if index < 1 or index > len(candidates):
                    valid = False
                    break
                chosen.append(candidates[index - 1][1])
            if not valid:
                print("Geçersiz seçim. Lütfen listedeki numaraları kullanın.")
                continue
            selection = chosen
            break
        if raw.isdigit():
            index = int(raw)
            if 1 <= index <= len(candidates):
                selection = [candidates[index - 1][1]]
                break
        print("Geçersiz seçim. Lütfen listedeki numaraları kullanın.")

    return selection


def determine_column_mapping(
    master_df: pd.DataFrame,
    definitions: Sequence[FieldDefinition],
    mapping_path: Optional[Path],
    logger: logging.Logger,
    non_interactive: bool,
) -> Dict[str, List[str]]:
    """Determine the best column matches for each ``FieldDefinition``."""

    saved_mappings = _load_saved_mappings(mapping_path)
    resolved: Dict[str, List[str]] = {}
    should_save = False

    master_columns = [col for col in master_df.columns if isinstance(col, str) and col.strip()]

    for definition in definitions:
        existing = [col for col in saved_mappings.get(definition.canonical, []) if col in master_columns]
        if existing:
            resolved[definition.canonical] = existing
            continue

        scored_candidates: List[Tuple[float, str]] = []
        for column in master_columns:
            score = _compute_match_score(column, definition.synonyms)
            if score > 0:
                scored_candidates.append((score, column))

        scored_candidates.sort(reverse=True, key=lambda item: item[0])

        chosen: List[str] = []
        if definition.multi_source:
            chosen = [name for score, name in scored_candidates if score >= MULTI_ACCEPT_THRESHOLD][: definition.max_items]
            if not chosen and scored_candidates:
                if non_interactive or not sys.stdin.isatty():
                    logger.info(
                        "'%s' alanı için güvenilir eşleşme bulunamadı; atlanıyor",
                        definition.description,
                    )
                    continue
                chosen = _prompt_for_selection(definition, scored_candidates[:10], allow_multiple=True)
        else:
            if scored_candidates and scored_candidates[0][0] >= AUTO_ACCEPT_THRESHOLD:
                chosen = [scored_candidates[0][1]]
            elif scored_candidates and scored_candidates[0][0] >= PROMPT_THRESHOLD:
                if non_interactive or not sys.stdin.isatty():
                    logger.info(
                        "'%s' alanı için düşük skorlu eşleşme '%s' kullanılıyor (skor %.2f)",
                        definition.description,
                        scored_candidates[0][1],
                        scored_candidates[0][0],
                    )
                    chosen = [scored_candidates[0][1]]
                else:
                    chosen = _prompt_for_selection(definition, scored_candidates[:10], allow_multiple=False)
            elif scored_candidates:
                if non_interactive or not sys.stdin.isatty():
                    logger.info(
                        "'%s' alanı için eşleşme bulunamadı; en iyi aday '%s' (skor %.2f) fakat kullanıcı onayı gerekli",
                        definition.description,
                        scored_candidates[0][1],
                        scored_candidates[0][0],
                    )
                    continue
                chosen = _prompt_for_selection(definition, scored_candidates[:10], allow_multiple=False)

        if chosen:
            resolved[definition.canonical] = chosen[: definition.max_items]
            should_save = True
            continue

        logger.info("Kaynak dosyada '%s' alanı için eşleşen kolon bulunamadı", definition.description)

    if should_save:
        _save_mappings(mapping_path, resolved)

    return resolved


# ---------------------------------------------------------------------------
# Mapping logic
# ---------------------------------------------------------------------------


def _extract_values(row: Mapping[str, object], sources: Sequence[str], definition: FieldDefinition) -> List[str]:
    values: List[str] = []
    for column in sources:
        raw = row.get(column)
        cleaned = _clean_value(raw)
        if not cleaned:
            continue
        if definition.split_semicolon:
            values.extend(_split_semicolon(cleaned, definition.max_items))
        else:
            values.append(cleaned)
    if definition.max_items is not None:
        return values[: definition.max_items]
    return values


def apply_mapping(
    master_df: pd.DataFrame,
    sheet: Worksheet,
    headers: Mapping[str, int],
    valid_map: Mapping[str, Iterable[str]],
    definitions: Sequence[FieldDefinition],
    column_mapping: Mapping[str, List[str]],
    logger: logging.Logger,
) -> Dict[str, int]:
    """Populate ``sheet`` with data from ``master_df`` using ``column_mapping``."""

    stats: Dict[str, int] = {"products": 0, "filled": 0, "skipped": 0, "normalized": 0}

    column_indices: Dict[str, int] = {}
    missing_targets: List[str] = []

    for definition in definitions:
        for target in definition.template_targets:
            idx = _match_column(headers, target)
            if idx is not None:
                column_indices[target] = idx
            else:
                missing_targets.append(target)

    prop65_column = _match_column(headers, PROP65_TARGET)
    if prop65_column is not None:
        column_indices[PROP65_TARGET] = prop65_column
    else:
        missing_targets.append(PROP65_TARGET)

    if missing_targets:
        for target in missing_targets:
            logger.info("Şablonda '%s' kolonu bulunamadı; atlanacak", target)

    cleaned_df = master_df.where(pd.notna(master_df), None)
    row_index = 2

    sku_sources = column_mapping.get("core::supplierPartNumber", [])
    if not sku_sources:
        logger.info("SKU eşleşmesi bulunamadı; ürün satırları atlanacak")
        return stats

    missing_source_logged: set[str] = set()

    sku_definition = FieldDefinition(
        canonical="core::supplierPartNumber",
        template_targets=["core::supplierPartNumber"],
        synonyms=[],
        description="SKU / RugNo",
    )

    for _, row in cleaned_df.iterrows():
        sku_values = _extract_values(row, sku_sources, sku_definition)
        sku = sku_values[0] if sku_values else ""
        if not sku:
            logger.info("SKU değeri olmayan satır atlandı")
            stats["skipped"] += 1
            continue

        stats["products"] += 1

        for definition in definitions:
            template_columns = [column_indices[target] for target in definition.template_targets if target in column_indices]
            if not template_columns:
                continue

            sources = column_mapping.get(definition.canonical, [])
            if not sources:
                if definition.canonical not in missing_source_logged:
                    logger.info("%s alanı için kaynak kolon bulunamadı", definition.description)
                    missing_source_logged.add(definition.canonical)
                continue

            values = _extract_values(row, sources, definition)
            for idx, column_index in enumerate(template_columns):
                value = values[idx] if idx < len(values) else ""
                normalized_value, was_normalized, _is_valid, had_input = normalize_with_valid_values(
                    definition.template_targets[min(idx, len(definition.template_targets) - 1)],
                    value,
                    valid_map,
                    logger,
                    alias=definition.valid_key,
                )
                sheet.cell(row=row_index, column=column_index, value=normalized_value or None)
                if normalized_value:
                    stats["filled"] += 1
                    if was_normalized:
                        stats["normalized"] += 1
                elif had_input:
                    stats["skipped"] += 1

        if prop65_column is not None:
            cell = sheet.cell(row=row_index, column=prop65_column)
            existing = _clean_value(cell.value)
            if not existing:
                sheet.cell(row=row_index, column=prop65_column, value="No")
            else:
                sheet.cell(row=row_index, column=prop65_column, value=existing)

        row_index += 1

    return stats


# ---------------------------------------------------------------------------
# Additional images sheet handling
# ---------------------------------------------------------------------------


def write_additional_images(
    workbook,
    master_df: pd.DataFrame,
    logger: logging.Logger,
    sku_sources: Sequence[str],
    image_sources: Sequence[str],
) -> Tuple[int, int]:
    """Populate the "Additional Images" worksheet."""

    if not sku_sources or not image_sources:
        logger.info("Ek görseller için gerekli kolonlar eşleşmedi; sayfa atlanıyor")
        return 0, 0

    sheet = _find_sheet(workbook, "Additional Images")
    if sheet is None:
        logger.info("'Additional Images' sayfası bulunamadı")
        return 0, 0

    headers = detect_template_columns(sheet)

    sku_column = _match_column(headers, "core::supplierPartNumber") or _match_column(headers, "Supplier Part Number")
    image_url_column = _match_column(headers, "Image URL") or _match_column(headers, "Image Url")
    is_primary_column = _match_column(headers, "Is Primary")
    sequence_column = _match_column(headers, "Sequence")

    if sku_column is None or image_url_column is None:
        logger.info("Ek görseller sayfasında zorunlu kolonlar yok; atlanıyor")
        return 0, 0

    next_row = sheet.max_row + 1 if sheet.max_row else 2
    written = 0
    skipped = 0

    cleaned_df = master_df.where(pd.notna(master_df), None)

    sku_definition = FieldDefinition(
        canonical="core::supplierPartNumber",
        template_targets=["core::supplierPartNumber"],
        synonyms=[],
        description="SKU / RugNo",
    )
    image_definition = FieldDefinition(
        canonical="media::imageUrls",
        template_targets=["media::image1"],
        synonyms=[],
        description="Image URLs",
        split_semicolon=True,
        max_items=50,
    )

    for _, row in cleaned_df.iterrows():
        sku_values = _extract_values(row, sku_sources, sku_definition)
        sku = sku_values[0] if sku_values else ""
        if not sku:
            continue

        image_values = _extract_values(row, image_sources, image_definition)
        urls = [url for url in image_values if url]
        if not urls:
            skipped += 1
            continue

        for order, url in enumerate(urls, start=1):
            sheet.cell(row=next_row, column=sku_column, value=sku)
            sheet.cell(row=next_row, column=image_url_column, value=url)
            if is_primary_column is not None:
                sheet.cell(row=next_row, column=is_primary_column, value="Yes" if order == 1 else "No")
            if sequence_column is not None:
                sheet.cell(row=next_row, column=sequence_column, value=order)
            next_row += 1
            written += 1

    if written == 0 and skipped == 0:
        logger.info("Ek görsel URL'leri yazılmadı")

    return written, skipped


# ---------------------------------------------------------------------------
# Workbook saving
# ---------------------------------------------------------------------------


def save_filled_workbook(workbook, template_path: Path, output_path: Optional[Path] = None) -> Path:
    """Save ``workbook`` to ``output_path`` (defaults to ``*_filled.xlsx``)."""

    if output_path is None:
        output_path = template_path.with_name(f"{template_path.stem}_filled{template_path.suffix}")

    workbook.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Data loading & template selection
# ---------------------------------------------------------------------------


def _load_master_dataframe(master_path: Path) -> pd.DataFrame:
    """Load the master dataset using pandas in a memory conscious manner."""

    suffix = master_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(master_path, dtype=str)

    if suffix in {".csv", ".txt"}:
        return pd.read_csv(master_path, dtype=str, keep_default_na=False)

    raise ValueError(f"Unsupported master file format: {master_path.suffix}")


def _select_main_sheet(workbook) -> Optional[Worksheet]:
    if not workbook.sheetnames:
        return None

    best_sheet = workbook[workbook.sheetnames[0]]
    best_count = len(detect_template_columns(best_sheet))

    for sheet_name in workbook.sheetnames[1:]:
        sheet = workbook[sheet_name]
        header_count = len(detect_template_columns(sheet))
        if header_count > best_count:
            best_sheet = sheet
            best_count = header_count

    return best_sheet


def _locate_default_template() -> Optional[Path]:
    candidates = [
        Path("wayfair_template.xlsx"),
        Path(__file__).with_name("wayfair_template.xlsx"),
        Path(__file__).parent / "templates" / "wayfair_template.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


# ---------------------------------------------------------------------------
# Main orchestration
# ---------------------------------------------------------------------------


def run_mapper(
    master_path: Path,
    template_path: Optional[Path],
    output_path: Optional[Path],
    mapping_path: Optional[Path],
    non_interactive: bool,
) -> None:
    """Run the full mapping pipeline."""

    log_path = Path("mapper.log")
    logger = setup_logger(log_path)

    logger.info("Wayfair mapper başlatılıyor")
    logger.info("Master veri: %s", master_path)

    actual_template_path = template_path or _locate_default_template()
    if actual_template_path is None:
        logger.error("Kullanılabilir bir Wayfair şablonu bulunamadı")
        print("Error: Wayfair template not found. See mapper.log for details.")
        return

    logger.info("Şablon: %s", actual_template_path)

    try:
        master_df = _load_master_dataframe(master_path)
    except Exception as exc:  # pragma: no cover - defensive logging
        logger.exception("Master verisi okunamadı: %s", exc)
        print("Error: Unable to read master data. See mapper.log for details.")
        return

    try:
        workbook = load_workbook(actual_template_path)
    except Exception as exc:  # pragma: no cover - defensive logging
        logger.exception("Şablon dosyası açılamadı: %s", exc)
        print("Error: Unable to load template workbook. See mapper.log for details.")
        return

    main_sheet = _select_main_sheet(workbook)
    if main_sheet is None:
        logger.error("Şablon içinde veri sayfası bulunamadı")
        print("Error: Main template sheet missing. See mapper.log for details.")
        return

    headers = detect_template_columns(main_sheet)
    if not headers:
        logger.error("Şablonda kolon başlıkları okunamadı")
        print("Error: Failed to read template headers. See mapper.log for details.")
        return

    valid_map = _load_valid_values(workbook)
    if valid_map:
        logger.info("'Valid Values' sayfasından %s alan okundu", len(valid_map))
    else:
        logger.info("'Valid Values' sayfası bulunamadı veya boş")

    column_mapping = determine_column_mapping(master_df, FIELD_DEFINITIONS, mapping_path, logger, non_interactive)

    stats = apply_mapping(master_df, main_sheet, headers, valid_map, FIELD_DEFINITIONS, column_mapping, logger)

    sku_sources = column_mapping.get("core::supplierPartNumber", [])
    image_sources = column_mapping.get("media::imageUrls", [])
    additional_written, additional_skipped = write_additional_images(
        workbook, master_df, logger, sku_sources, image_sources
    )

    output = save_filled_workbook(workbook, actual_template_path, output_path)

    logger.info("Dosya kaydedildi: %s", output)
    logger.info("Ürün sayısı: %s", stats.get("products", 0))
    logger.info("Dolu alanlar: %s", stats.get("filled", 0))
    logger.info("Normalize edilen değerler: %s", stats.get("normalized", 0))
    logger.info("Atlanan değerler: %s", stats.get("skipped", 0))
    logger.info("Ek görsel satırları: %s", additional_written)
    logger.info("Ek görsel atlamaları: %s", additional_skipped)

    summary = (
        f"Products processed: {stats.get('products', 0)} | "
        f"Fields filled: {stats.get('filled', 0)} | "
        f"Normalized: {stats.get('normalized', 0)} | "
        f"Skipped: {stats.get('skipped', 0)} | "
        f"Additional image rows: {additional_written} | "
        f"Additional image skips: {additional_skipped}"
    )
    print(summary)


def _parse_args(argv: Optional[Iterable[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill a Wayfair Product Addition Template with master data.")
    parser.add_argument("--input", required=True, type=Path, help="Path to the master data file (Excel or CSV)")
    parser.add_argument("--out", type=Path, default=None, help="Optional output path for the filled workbook")
    parser.add_argument(
        "--template",
        type=Path,
        default=None,
        help="Path to the Wayfair template workbook. If omitted, attempts to use the bundled template.",
    )
    parser.add_argument(
        "--mappings",
        type=Path,
        default=None,
        help="Optional YAML file to load/save column mapping preferences.",
    )
    parser.add_argument(
        "--non-interactive",
        action="store_true",
        help="Do not prompt for column mapping confirmations; use best matches automatically.",
    )
    return parser.parse_args(argv)


def main(argv: Optional[Iterable[str]] = None) -> None:
    args = _parse_args(argv)
    run_mapper(args.input, args.template, args.out, args.mappings, args.non_interactive)


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    main(sys.argv[1:])
